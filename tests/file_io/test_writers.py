"""Tests for CSV and XLSX writers: basic writes, overwrite, atomic cleanup, summaries."""

import csv
import os
import time
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from openstaad_mcp.file_io import CSVWriter, write_output_file
from openstaad_mcp.file_io.path_validator import FileIOError

# ═══════════════════════════════════════════════════════════════════════════
# CSV Write
# ═══════════════════════════════════════════════════════════════════════════


class TestCSVWrite:
    def test_basic_write(self, tmp_path: Path):
        p = tmp_path / "out.csv"
        data = [["Member", "Fx"], [1, -12.4], [2, -8.1]]
        write_output_file(str(p), data, allowed_dirs=[tmp_path])
        assert p.exists()
        with open(p, encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert rows[0] == ["Member", "Fx"]

    def test_file_exists_no_overwrite(self, tmp_path: Path):
        p = tmp_path / "exists.csv"
        p.write_text("old", encoding="utf-8")
        with pytest.raises(FileIOError) as exc_info:
            write_output_file(str(p), [["a"], [1]], allowed_dirs=[tmp_path], overwrite=False)
        assert exc_info.value.code == "FILE_EXISTS"

    def test_overwrite_true(self, tmp_path: Path):
        p = tmp_path / "exists.csv"
        p.write_text("old", encoding="utf-8")
        write_output_file(str(p), [["a"], [1]], allowed_dirs=[tmp_path], overwrite=True)
        assert "a" in p.read_text(encoding="utf-8")

    def test_atomic_write_cleanup(self, tmp_path: Path):
        """No temp files left after successful write."""
        p = tmp_path / "out.csv"
        write_output_file(str(p), [["x"], [1]], allowed_dirs=[tmp_path])
        temps = list(tmp_path.glob(".~omcp_*"))
        assert temps == []

    def test_stale_temp_cleanup(self, tmp_path: Path):
        """Old temp files are cleaned up before a new write."""
        stale = tmp_path / ".~omcp_stale.tmp"
        stale.write_text("orphan", encoding="utf-8")
        old_time = time.time() - 7200
        os.utime(stale, (old_time, old_time))

        p = tmp_path / "out.csv"
        write_output_file(str(p), [["x"], [1]], allowed_dirs=[tmp_path])
        assert not stale.exists()


# ═══════════════════════════════════════════════════════════════════════════
# XLSX Write
# ═══════════════════════════════════════════════════════════════════════════


class TestXLSXWrite:
    def test_flat_write(self, tmp_path: Path):
        p = tmp_path / "out.xlsx"
        data = [["Member", "Fx"], [1, -12.4]]
        write_output_file(str(p), data, allowed_dirs=[tmp_path])
        wb = openpyxl.load_workbook(p)
        ws = wb.active
        assert isinstance(ws, Worksheet)
        assert ws.cell(1, 1).value == "Member"
        assert ws.cell(2, 2).value == -12.4

    def test_multi_sheet_write(self, tmp_path: Path):
        p = tmp_path / "multi.xlsx"
        data = {
            "Summary": {"columns": ["Member", "Status"], "rows": [[1, "OK"]]},
            "Details": {"columns": ["Member", "Fx"], "rows": [[1, -12.4]]},
        }
        write_output_file(str(p), data, allowed_dirs=[tmp_path])
        wb = openpyxl.load_workbook(p)
        assert "Summary" in wb.sheetnames
        assert "Details" in wb.sheetnames

    def test_overwrite_existing_xlsx(self, tmp_path: Path):
        """Overwriting an existing XLSX file replaces the content entirely."""
        p = tmp_path / "overwrite.xlsx"
        write_output_file(str(p), [["old_col"], ["old_val"]], allowed_dirs=[tmp_path])
        write_output_file(str(p), [["new_col"], ["new_val"]], allowed_dirs=[tmp_path], overwrite=True)
        wb = openpyxl.load_workbook(p)
        ws = wb.active
        assert isinstance(ws, Worksheet)
        assert ws.cell(1, 1).value == "new_col"
        assert ws.cell(2, 1).value == "new_val"

    def test_multi_sheet_overwrite_preserves_new_sheets(self, tmp_path: Path):
        """Overwriting with different sheet names produces only the new sheets."""
        p = tmp_path / "sheets.xlsx"
        old_data = {"OldSheet": {"columns": ["a"], "rows": [[1]]}}
        new_data = {"NewSheet": {"columns": ["b"], "rows": [[2]]}}
        write_output_file(str(p), old_data, allowed_dirs=[tmp_path])
        write_output_file(str(p), new_data, allowed_dirs=[tmp_path], overwrite=True)
        wb = openpyxl.load_workbook(p)
        assert "NewSheet" in wb.sheetnames
        assert "OldSheet" not in wb.sheetnames


# ═══════════════════════════════════════════════════════════════════════════
# Output summary
# ═══════════════════════════════════════════════════════════════════════════


class TestOutputSummary:
    def test_output_summary_csv(self, tmp_path: Path):
        p = tmp_path / "out.csv"
        data = [["Member", "Fx"], [1, -12.4], [2, -8.1]]
        summary = CSVWriter(p).build_summary(data)
        assert summary["rows_written"] == 2
        assert summary["columns"] == ["Member", "Fx"]
        assert len(summary["sample_rows"]) == 2
