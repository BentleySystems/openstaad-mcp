"""
Server-side file I/O for the ``execute_code`` tool.

All operations run **outside** the sandbox.  The sandbox never touches the
filesystem — it receives pre-parsed data via ``__input__`` and returns a
structured value that this module writes to disk.

Architecture
------------
Reading and writing are handled by format-specific subclasses:

- :class:`CSVReader` / :class:`CSVWriter`
- :class:`XLSXReader` / :class:`XLSXWriter`

Each inherits from :class:`BaseReader` or :class:`BaseWriter`, which enforce
file-size limits, column/row caps, and atomic writes.

Public API
----------
``read_input_file``  /  ``write_output_file``
    Dispatch to the correct reader/writer based on file extension.

``get_allowed_dirs``  /  ``get_input_data``
    Server-level helpers called from ``execute_code``.

``validate_return_value``  /  ``deep_freeze``
    Data validation and summary helpers.
"""

from __future__ import annotations

import abc
import csv
import logging
import os
import time
import uuid
from datetime import date, datetime
from datetime import time as dt_time
from pathlib import Path
from typing import Any

import chardet
import openpyxl
from fastmcp.server.context import Context
from mcp.shared.exceptions import McpError

from openstaad_mcp.sandbox.const import (
    MAX_FILE_SIZE_BYTES,
    MAX_INPUT_COLUMNS,
    MAX_INPUT_ROWS,
    MAX_INPUT_SHEETS,
    MAX_OUTPUT_COLUMNS,
    MAX_OUTPUT_ROWS,
    MAX_OUTPUT_SHEETS,
    MAX_SHEET_NAME_LENGTH,
    SAMPLE_ROW_COUNT,
    STALE_TEMP_AGE_SECONDS,
    TEMP_FILE_PREFIX,
)
from openstaad_mcp.file_io.path_validator import FileIOError, parse_roots_to_dirs, validate_io_path

logger = logging.getLogger(__name__)

_JSON_PRIMITIVES = (str, int, float, bool, type(None))


# ═══════════════════════════════════════════════════════════════════════════
# Server-level helpers (called from server.py)
# ═══════════════════════════════════════════════════════════════════════════


def validate_args_allowed_dirs(allowed_dirs: list[str] | None) -> list[Path]:
    """Validate and resolve ``--allowed-dir`` CLI arguments to real paths.

    Security: resolves symlinks so later checks compare against real paths.
    """
    if not allowed_dirs:
        return []

    result: list[Path] = []
    for dir_str in allowed_dirs:
        # Expand ~/… to the user's home directory
        expanded = Path(dir_str).expanduser()
        absolute = expanded.resolve(strict=False)
        normalized_original = Path(os.path.normpath(absolute))

        try:
            # Security: resolve symlinks in allowed directories during startup
            resolved = absolute.resolve(strict=True)
            normalized_resolved = Path(os.path.normpath(resolved))
            result.append(normalized_resolved)
        except OSError:
            # If we can't resolve (doesn't exist), use the normalized absolute path
            # This allows configuring allowed dirs that will be created later
            result.append(normalized_original)

    return result


async def get_allowed_dirs(
    ctx: Context, args_allowed_dirs: list[Path], input_path: str | None, output_path: str | None
) -> list[Path]:
    """Resolve MCP roots into a list of allowed directories."""
    logger.debug("Args allowed dirs: %s", args_allowed_dirs)
    allowed_dirs: list[Path] = [Path(el) for el in args_allowed_dirs]
    if input_path is not None or output_path is not None:
        try:
            roots = await ctx.list_roots()
            logger.debug(f"Received MCP roots: {roots}")
        except McpError as exc:
            logger.error(f"Error listing MCP roots: {exc}")
            roots = []
        allowed_dirs += parse_roots_to_dirs(roots)
    logger.debug(f"Allowed directories for file I/O: {allowed_dirs}")
    return allowed_dirs


async def get_input_data(input_path: str | None, allowed_dirs: list[Path]) -> tuple[Any, dict[str, Any] | None]:
    """Validate path, read file, freeze data.  Returns ``(data, summary)``."""
    if input_path is None:
        return None, None
    resolved_input = validate_io_path(input_path, allowed_dirs, mode="read")
    data, input_summary = read_input_file(resolved_input)
    return deep_freeze(data), input_summary


# ═══════════════════════════════════════════════════════════════════════════
# Base reader
# ═══════════════════════════════════════════════════════════════════════════


class BaseReader(abc.ABC):
    """Base class for file readers.  Enforces file-size and limit checks."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self._check_file_size()

    def _check_file_size(self) -> None:
        size = self.path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            raise FileIOError(
                "FILE_TOO_LARGE",
                f"File is {size:,} bytes; limit is {MAX_FILE_SIZE_BYTES:,} bytes",
            )

    @abc.abstractmethod
    def read(self, *, start_row: int = 0, max_rows: int | None = None, **kwargs: Any) -> Any:
        """Parse the file and return structured data."""

    @abc.abstractmethod
    def build_summary(self, data: Any) -> dict[str, Any]:
        """Build a lightweight summary for the agent."""


# ── Header detection ─────────────────────────────────────────────────────


def _cell_type(value: Any) -> str:
    """Classify a cell value for header-detection comparison."""
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "numeric"
    if isinstance(value, (datetime, date, dt_time)):
        return "numeric"
    return "string"


def _detect_header(rows: list[list], has_header: bool | None) -> bool:
    """Detect whether the first row is a header.

    When *has_header* is ``None`` (auto-detect), samples up to 5 rows and
    compares the per-column type of row 0 against the majority type of
    rows 1-4.  If any column's first-row type differs from its data-row
    majority, the first row is treated as a header.
    """
    if has_header is not None:
        return has_header

    if len(rows) <= 1:
        return True  # Too few rows to compare — conservative default

    first_row = rows[0]
    data_rows = rows[1:5]  # Up to 4 data rows for comparison

    if not first_row:
        return True

    for col_idx in range(len(first_row)):
        first_type = _cell_type(first_row[col_idx])
        if first_type == "null":
            continue

        type_counts: dict[str, int] = {}
        for row in data_rows:
            if col_idx < len(row):
                t = _cell_type(row[col_idx])
                if t != "null":
                    type_counts[t] = type_counts.get(t, 0) + 1

        if not type_counts:
            continue

        majority_type = max(type_counts, key=type_counts.get)
        if first_type != majority_type:
            return True  # Type mismatch → first row is a header

    return False  # All columns match → not a header


def _auto_columns(num_cols: int) -> list[str]:
    """Generate column names ``col_1, col_2, …`` for headerless data."""
    return [f"col_{i + 1}" for i in range(num_cols)]


# ── CSV reader ───────────────────────────────────────────────────────────


class CSVReader(BaseReader):
    """Reads a CSV file into ``list[list]`` (array-of-arrays).

    Uses ``chardet`` for encoding detection and ``csv.Sniffer`` for
    dialect detection.  Values are coerced from strings to int/float
    where possible.  Streams from disk line-by-line.
    """

    _CHARDET_MIN_CONFIDENCE = 0.5

    def read(
        self, *, start_row: int = 0, max_rows: int | None = None, has_header: bool | None = None, **kwargs: Any
    ) -> list[list]:
        encoding = self._detect_encoding()
        dialect = self._detect_dialect(encoding)
        all_rows: list[list] = []

        with open(self.path, newline="", encoding=encoding) as f:
            reader = csv.reader(f, dialect)
            for raw_row in reader:
                coerced = [_coerce_csv_value(v) for v in raw_row]
                if len(coerced) > MAX_INPUT_COLUMNS:
                    raise FileIOError(
                        "TOO_MANY_COLUMNS",
                        f"Row has {len(coerced)} columns; limit is {MAX_INPUT_COLUMNS}",
                    )
                all_rows.append(coerced)

        self._has_header = _detect_header(all_rows, has_header)

        if self._has_header:
            if not all_rows:
                return []
            data_rows = all_rows[1:]
            if len(data_rows) > MAX_INPUT_ROWS:
                raise FileIOError(
                    "TOO_MANY_ROWS",
                    f"File has {len(data_rows)} data rows; limit is {MAX_INPUT_ROWS}",
                )
            sliced = data_rows[start_row:]
            if max_rows is not None:
                sliced = sliced[:max_rows]
            return [all_rows[0], *sliced] if start_row == 0 else sliced
        else:
            if len(all_rows) > MAX_INPUT_ROWS:
                raise FileIOError(
                    "TOO_MANY_ROWS",
                    f"File has {len(all_rows)} data rows; limit is {MAX_INPUT_ROWS}",
                )
            sliced = all_rows[start_row:]
            if max_rows is not None:
                sliced = sliced[:max_rows]
            return sliced

    def build_summary(self, data: list[list]) -> dict[str, Any]:
        has_header = getattr(self, "_has_header", True)
        if has_header:
            header = data[0] if data else []
            data_rows = data[1:] if len(data) > 1 else []
        else:
            num_cols = len(data[0]) if data else 0
            header = _auto_columns(num_cols)
            data_rows = data
        return {
            "total_rows": len(data_rows),
            "columns": list(header),
            "sample_rows": [list(r) for r in data_rows[:SAMPLE_ROW_COUNT]],
        }

    # -- helpers --

    def _detect_encoding(self) -> str:
        """Detect file encoding: try UTF-8 first, then chardet, then cp1252."""
        raw = self.path.read_bytes()
        result = chardet.detect(raw)
        encoding = result.get("encoding")
        confidence = result.get("confidence", 0)
        if encoding and confidence >= self._CHARDET_MIN_CONFIDENCE:
            return encoding
        # Low confidence or no result — fall back to cp1252 (Windows default)
        return "cp1252"

    def _detect_dialect(self, encoding: str) -> type[csv.Dialect]:
        """Detect CSV dialect using csv.Sniffer, fall back to ``excel``.

        Only trusts the sniffer when the detected delimiter is a common
        separator character.  Exotic delimiters (letters, digits, etc.)
        indicate a false positive from a small or ambiguous sample.
        """
        _COMMON_DELIMITERS = {",", ";", "\t", "|"}
        try:
            with open(self.path, newline="", encoding=encoding) as f:
                sample = f.read(8192)
            dialect = csv.Sniffer().sniff(sample)
            if dialect.delimiter in _COMMON_DELIMITERS:
                return dialect
        except csv.Error:
            pass
        return csv.excel


def _coerce_csv_value(val: str) -> int | float | str:
    """Attempt int → float → str coercion of a CSV string value."""
    try:
        return int(val)
    except ValueError:
        pass
    try:
        return float(val)
    except ValueError:
        pass
    return val


# ── XLSX reader ──────────────────────────────────────────────────────────


class XLSXReader(BaseReader):
    """Reads an XLSX workbook into ``{sheet_name: {columns, rows}}``."""

    def read(
        self,
        *,
        start_row: int = 0,
        max_rows: int | None = None,
        sheet: str | None = None,
        has_header: bool | None = None,
        **kwargs: Any,
    ) -> dict[str, dict[str, Any]]:
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        except Exception as exc:
            raise FileIOError("CORRUPTED_WORKBOOK", f"Cannot open workbook: {exc}") from None

        try:
            self._validate_sheet_count(wb)
            sheets_to_load = self._resolve_sheets(wb, sheet)
            return {name: self._read_sheet(wb[name], name, start_row, max_rows, has_header) for name in sheets_to_load}
        finally:
            wb.close()

    def build_summary(self, data: dict[str, dict[str, Any]]) -> dict[str, Any]:
        sheets = list(data.keys())
        first_sheet = sheets[0] if sheets else None
        first = data[first_sheet] if first_sheet else {"columns": [], "rows": []}
        return {
            "sheets": sheets,
            "loaded_sheet": first_sheet,
            "total_rows": len(first["rows"]),
            "columns": list(first["columns"]),
            "sample_rows": [list(r) for r in first["rows"][:SAMPLE_ROW_COUNT]],
        }

    # -- helpers --

    @staticmethod
    def _validate_sheet_count(wb: Any) -> None:
        if len(wb.sheetnames) > MAX_INPUT_SHEETS:
            raise FileIOError(
                "TOO_MANY_ROWS",
                f"Workbook has {len(wb.sheetnames)} sheets; limit is {MAX_INPUT_SHEETS}",
            )

    @staticmethod
    def _resolve_sheets(wb: Any, sheet: str | None) -> list[str]:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise FileIOError("SHEET_NOT_FOUND", f"Sheet '{sheet}' not found in workbook")
            return [sheet]
        return list(wb.sheetnames)

    @staticmethod
    def _read_sheet(
        ws: Any, name: str, start_row: int, max_rows: int | None, has_header: bool | None
    ) -> dict[str, Any]:
        raw_rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(list(row))

        # Check column count on the first row
        if raw_rows and len(raw_rows[0]) > MAX_INPUT_COLUMNS:
            raise FileIOError(
                "TOO_MANY_COLUMNS",
                f"Sheet '{name}' has {len(raw_rows[0])} columns; limit is {MAX_INPUT_COLUMNS}",
            )

        # Detect header using raw types (before datetime → string conversion)
        is_header = _detect_header(raw_rows, has_header)

        # Convert to JSON primitives
        all_rows = [[_to_json_primitive(c) for c in row] for row in raw_rows]

        if is_header:
            columns = all_rows[0] if all_rows else []
            data_rows = all_rows[1:]
        else:
            num_cols = len(all_rows[0]) if all_rows else 0
            columns = _auto_columns(num_cols)
            data_rows = all_rows

        if len(data_rows) > MAX_INPUT_ROWS:
            raise FileIOError(
                "TOO_MANY_ROWS",
                f"Sheet '{name}' exceeds {MAX_INPUT_ROWS} rows",
            )

        sliced = data_rows[start_row:]
        if max_rows is not None:
            sliced = sliced[:max_rows]

        return {"columns": columns, "rows": sliced}


def _to_json_primitive(value: Any) -> str | int | float | bool | None:
    """Convert an openpyxl cell value to a JSON-safe primitive."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dt_time):
        return value.isoformat()
    return str(value)


# ═══════════════════════════════════════════════════════════════════════════
# Base writer
# ═══════════════════════════════════════════════════════════════════════════


class BaseWriter(abc.ABC):
    """Base class for file writers.  Handles atomic writes via temp file."""

    def __init__(self, path: Path) -> None:
        self.path = path

    def write(self, data: Any, *, overwrite: bool = False) -> dict[str, Any]:
        """Validate, write atomically, and return a summary."""
        if self.path.exists() and not overwrite:
            raise FileIOError("FILE_EXISTS", f"File already exists: {self.path}")
        _clean_stale_temps(self.path.parent)

        tmp = _temp_path(self.path.parent)
        try:
            self._write_to(tmp, data)
            os.replace(tmp, self.path)
        except BaseException:
            tmp.unlink(missing_ok=True)
            raise
        return self.build_summary(data)

    @abc.abstractmethod
    def _write_to(self, tmp: Path, data: Any) -> None:
        """Write *data* to the temporary file *tmp*."""

    @abc.abstractmethod
    def build_summary(self, data: Any) -> dict[str, Any]:
        """Build a lightweight summary for the agent."""


# ── CSV writer ───────────────────────────────────────────────────────────


class CSVWriter(BaseWriter):
    """Writes ``list[list]`` to a CSV file."""

    def _write_to(self, tmp: Path, data: list[list]) -> None:
        with open(tmp, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(data)

    def build_summary(self, data: list[list]) -> dict[str, Any]:
        header = data[0] if data else []
        data_rows = data[1:] if len(data) > 1 else []
        return {
            "message": f"The `result` data has been written to `{self.path}`",
            "rows_written": len(data_rows),
            "columns": list(header),
            "sample_rows": [list(r) for r in data_rows[:SAMPLE_ROW_COUNT]],
        }


# ── XLSX writer ──────────────────────────────────────────────────────────


class XLSXWriter(BaseWriter):
    """Writes flat or multi-sheet data to an XLSX file."""

    def _write_to(self, tmp: Path, data: Any) -> None:
        wb = openpyxl.Workbook()
        if isinstance(data, dict):
            for i, (name, sheet_data) in enumerate(data.items()):
                if i == 0:
                    ws = wb.active
                    assert ws is not None
                    ws.title = name
                else:
                    ws = wb.create_sheet(title=name)
                ws.append(sheet_data["columns"])
                for row in sheet_data["rows"]:
                    ws.append(row)
        else:
            ws = wb.active
            assert ws is not None
            for row in data:
                ws.append(row)
        wb.save(tmp)

    def build_summary(self, data: Any) -> dict[str, Any]:
        if isinstance(data, dict):
            return {
                "message": f"The `result` data has been written to `{self.path}`",
                "sheets": {
                    name: {
                        "columns": sheet["columns"],
                        "rows_written": len(sheet["rows"]),
                        "sample_rows": [list(r) for r in sheet["rows"][:SAMPLE_ROW_COUNT]],
                    }
                    for name, sheet in data.items()
                },
            }
        header = data[0] if data else []
        data_rows = data[1:] if len(data) > 1 else []
        return {
            "message": f"The `result` data has been written to `{self.path}`",
            "rows_written": len(data_rows),
            "columns": list(header),
            "sample_rows": [list(r) for r in data_rows[:SAMPLE_ROW_COUNT]],
        }


# ═══════════════════════════════════════════════════════════════════════════
# Reader/writer factory helpers
# ═══════════════════════════════════════════════════════════════════════════

_READERS: dict[str, type[BaseReader]] = {".csv": CSVReader, ".xlsx": XLSXReader}
_WRITERS: dict[str, type[BaseWriter]] = {".csv": CSVWriter, ".xlsx": XLSXWriter}


def _get_reader(path: Path) -> BaseReader:
    ext = path.suffix.lower()
    cls = _READERS.get(ext)
    if cls is None:
        raise FileIOError("UNSUPPORTED_FORMAT", f"Cannot read '{ext}' files")
    return cls(path)


def _get_writer(path: Path) -> BaseWriter:
    ext = path.suffix.lower()
    cls = _WRITERS.get(ext)
    if cls is None:
        raise FileIOError("UNSUPPORTED_FORMAT", f"Cannot write '{ext}' files")
    return cls(path)


# ═══════════════════════════════════════════════════════════════════════════
# Public dispatch functions (preserve existing API)
# ═══════════════════════════════════════════════════════════════════════════


def read_input_file(
    path: Path,
    *,
    sheet: str | None = None,
    start_row: int = 0,
    max_rows: int | None = None,
    has_header: bool | None = None,
) -> tuple[Any, dict[str, Any]]:
    """Read a CSV or XLSX file and return ``(data, summary)``."""
    reader = _get_reader(path)
    data = reader.read(start_row=start_row, max_rows=max_rows, sheet=sheet, has_header=has_header)
    summary = reader.build_summary(data)
    return data, summary


def write_output_file(path: str, data: Any, allowed_dirs: list[Path], *, overwrite: bool = False) -> dict[str, Any]:
    """Validate path, validate return value, write atomically, return summary."""
    resolved = validate_io_path(path, allowed_dirs, mode="write")
    validate_return_value(data)
    writer = _get_writer(resolved)
    return writer.write(data, overwrite=overwrite)


# ═══════════════════════════════════════════════════════════════════════════
# Return value validation
# ═══════════════════════════════════════════════════════════════════════════


def validate_return_value(value: Any) -> None:
    """Raise :class:`FileIOError` if *value* is not a valid output structure.

    Accepts:
    - ``list[list[primitive]]``  (flat / CSV / single-sheet)
    - ``dict[str, {columns: list, rows: list[list[primitive]]}]``  (multi-sheet)
    """
    if isinstance(value, list):
        _validate_flat(value)
    elif isinstance(value, dict):
        _validate_multi_sheet(value)
    else:
        raise FileIOError(
            "INVALID_RETURN_SHAPE",
            "Return value must be a list of lists (flat) or a dict of sheets (multi-sheet)",
        )


def _validate_flat(rows: list) -> None:
    if len(rows) > MAX_OUTPUT_ROWS + 1:  # +1 header
        raise FileIOError("INVALID_RETURN_SHAPE", f"Too many rows: {len(rows)}; limit {MAX_OUTPUT_ROWS}")
    for row in rows:
        if not isinstance(row, (list, tuple)):
            raise FileIOError("INVALID_RETURN_SHAPE", f"Each row must be a list, got {type(row).__name__}")
        if len(row) > MAX_OUTPUT_COLUMNS:
            raise FileIOError("INVALID_RETURN_SHAPE", f"Too many columns: {len(row)}; limit {MAX_OUTPUT_COLUMNS}")
        for cell in row:
            if not isinstance(cell, _JSON_PRIMITIVES):
                raise FileIOError(
                    "INVALID_RETURN_SHAPE",
                    f"Cell value must be a JSON primitive, got {type(cell).__name__}",
                )


def _validate_multi_sheet(sheets: dict) -> None:
    if len(sheets) > MAX_OUTPUT_SHEETS:
        raise FileIOError("INVALID_RETURN_SHAPE", f"Too many sheets: {len(sheets)}; limit {MAX_OUTPUT_SHEETS}")
    for name, sheet_data in sheets.items():
        if len(name) > MAX_SHEET_NAME_LENGTH:
            raise FileIOError(
                "INVALID_RETURN_SHAPE",
                f"Sheet name '{name}' exceeds {MAX_SHEET_NAME_LENGTH} characters",
            )
        if not isinstance(sheet_data, dict) or "columns" not in sheet_data or "rows" not in sheet_data:
            raise FileIOError(
                "INVALID_RETURN_SHAPE",
                "Each sheet must have 'columns' and 'rows' keys",
            )
        _validate_flat([sheet_data["columns"], *list(sheet_data["rows"])])


# ═══════════════════════════════════════════════════════════════════════════
# Deep freeze
# ═══════════════════════════════════════════════════════════════════════════


def deep_freeze(data: Any) -> Any:
    """Recursively convert mutable containers to immutable equivalents.

    - ``list`` → ``tuple``
    - ``dict`` values are recursively frozen (dict keys stay as-is since
      strings are already immutable)
    - Primitives (str, int, float, bool, None) pass through unchanged.
    """
    if data is None or isinstance(data, (str, int, float, bool)):
        return data
    if isinstance(data, (list, tuple)):
        return tuple(deep_freeze(item) for item in data)
    if isinstance(data, dict):
        return {k: deep_freeze(v) for k, v in data.items()}
    return data


# ═══════════════════════════════════════════════════════════════════════════
# Shared utilities
# ═══════════════════════════════════════════════════════════════════════════


def _temp_path(directory: Path) -> Path:
    return directory / f"{TEMP_FILE_PREFIX}{uuid.uuid4().hex}.tmp"


def _clean_stale_temps(directory: Path) -> None:
    """Remove orphaned temp files older than ``STALE_TEMP_AGE_SECONDS``."""
    cutoff = time.time() - STALE_TEMP_AGE_SECONDS
    for p in directory.glob(f"{TEMP_FILE_PREFIX}*.tmp"):
        try:
            if p.stat().st_mtime < cutoff:
                p.unlink()
        except OSError:
            pass
